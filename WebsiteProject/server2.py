#MAE website v2.1.2
#Made by Justin Fauson
import ctypes
import sys
import subprocess
import pyuac
import openpyxl
import pandas as pd

def is_admin():
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False

import time, os, csv, flask, random, webbrowser
import threading
#import VerificationPage
from flask import Flask, render_template, request, redirect, url_for


from Forms import SignInForm, FeedBackForm, AdminForm, SignUpForm, VerifyForm
from flask_wtf.csrf import CSRFProtect
from cryptography.fernet import Fernet
from datetime import datetime, timedelta
# Variables ---------------
emailList = '/Data/Email List FA25-2.csv'
loginTimeFile = '/Data/emails.csv'
logoutTimeFile = '/Data/Logout Times.csv'
feedbackFile = '/Data/feedback.csv'
pendingListFile = '/Data/pending list.csv'
AdminSettings = '/Admin/settings.xlsx'
eternalLoginPath = f'/Data/Logins-Eternal.csv'
eternalLogoutPath = f'/Data/Logins-Eternal.csv'
MasterPath = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__) #general convention for initializing flask app
csrf = CSRFProtect(app)
#random_int = random.randint(1231234213,2231234213)
app.config['SECRET_KEY'] = '1231234213'
app.config['WTF_CSRF_SECRET_KEY'] = '1231234213'
app.config['WTF_CSRF_TIME_LIMIT'] = 60*60*24*7
csrf.init_app(app)
end_time = 0.0
admin = False
#Sign in page

#-------Startup------
def LoadSettings():
    global emailListFile,loginTimeFile,logoutTimeFile,feedbackFile,pendingListFile
    AdminSettings = '/Admin/settings.xlsx'
    path = MasterPath
    settingsPath = path + AdminSettings
    wb = openpyxl.load_workbook(settingsPath)
    ws = wb.active
    semester = ws['B2'].value
    emailList = ws['B3'].value
    emailListFile = f'/Data/{emailList}'
    loginTimeFile = f'/Data/Logins-{semester}.csv'
    logoutTimeFile = f'/Data/Logouts-{semester}.csv'
    feedbackFile = '/Data/feedback.csv'
    pendingListFile = '/Data/pending list.csv'
    if not os.path.exists(path+loginTimeFile):
        with open(path+loginTimeFile, 'w', newline="") as f:
            writer = csv.writer(f)
            writer.writerow([f"Logins {semester}"])
    if not os.path.exists(path+logoutTimeFile):
        with open(path+logoutTimeFile, 'w', newline="") as f:
            writer = csv.writer(f)
            writer.writerow([f"Logins {semester}"])

#Webpages
@app.route('/', methods = ["get" ,"post"])
def signinPage():
    global end_time, globalemail, logoutTime,admin
    admin = False
    form = SignInForm()
    try:
        studentOfTheWeek = StudentOfTheWeek()
    except:
        studentOfTheWeek = "N/A"
    message = ""
    if form.is_submitted():
        result = request.form
        #Signin procedure
        if "signIn" in result:
            message,valid = signin(result)
            if valid:
                return redirect('/thankyou')
        #Signout procedure
        elif "Sign Out" in result or "altSignOut" in result:
            message,valid = signout(result)
            if valid:
                return redirect('/thankyouSignOut')
    message = ""
    #renders the webpage and sends the form variable to html
    return render_template('signin2.0.html', form=form, message=message, studentOfTheWeek=studentOfTheWeek)

def signin(result):
    message = ""
    global end_time, globalemail
    recordTime = time.time()
    inputEmail = result.get("email")+"@uccs.edu"
    #verification process
    pathR = MasterPath + emailList
    valid = False
    with open(pathR, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if row[0] == inputEmail:
                valid = True
                print(valid)
    pathW = MasterPath + loginTimeFile
    if valid:
        #write to csv file
        if not checkRecent(inputEmail, recordTime):
            with open(pathW, 'a', newline="") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow([inputEmail, recordTime])
        #print([result.get('email'), recordTime]) #display to console for testing
        # start timer for thank you page
        start_time = time.time()
        end_time = start_time + 1.0
        globalemail = inputEmail
    else:
        message = "Invalid Email"
    #return message
    print(3)
    return message, valid
    
def signout(result):
    global globalemail, logoutTime, end_time
    recordTime = time.time()
    message = ""
    # verification process
    inputEmail = result.get("email") + "@uccs.edu"
    pathR = MasterPath + emailList
    valid = False
    with open(pathR, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if row[0] == inputEmail:
                valid = True
                print(valid)
    #checkAccident(inputEmail,recordTime)#check if someone accidentaly signed in while signing out
    pathW = MasterPath + logoutTimeFile
    if valid:
        # write to csv file
        checkAccident(inputEmail,recordTime)
        with open(pathW, 'a', newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow([inputEmail, recordTime])
            print("logged out at",recordTime)  # display to console for testing
        print(result,"signed out")
        # start timer for thank you page
        start_time = time.time()
        end_time = start_time + 1.0
        logoutTime = recordTime
        globalemail = inputEmail
        
    else:
        message = "Invalid Email"
    return message, valid

def checkAccident(inputEmail,recordTime):
    GraceTime = float(recordTime)-120 #2 minutes to cancel logout
    pathR = MasterPath + loginTimeFile
    accident = False
    df = pd.read_csv(pathR)
    print(df)
    with open(pathR, 'r', newline="") as csvfile:
        lines = csvfile.readlines()
        last_line = lines[-1].split(",")
        lastEmail = last_line[0]
        print(lastEmail," test 1")
        lastTime = float(last_line[1].strip())
        print(lastTime, "test 2")
        if lastEmail == inputEmail:
            print("oof")
            if lastTime >= GraceTime:
                accident = True
        if accident:
            df = df.drop(df.index[-1])
            df.to_csv(pathR, index=False)

#Thank you page
@app.route('/thankyou')
def thankYou():
    global admin
    admin = False
    #if time is not up then render template
    if time.time()<=end_time:
        if globalemail != "":
            pathW = MasterPath + loginTimeFile
            count = 0
            with open(pathW, 'r', newline="") as csvfile:
                reader = csv.reader(csvfile)
                for row in reader:
                    if row[0] == globalemail:
                        count += 1
            message = "You have visited the MAE Help Center " + str(count) + " times!"
        else:
            message = ""
        return render_template('thankyou.html', message=message)
    #javascript reloads the page every second so that the redirect can happen
    return redirect('/')
@app.route('/thankyouSignOut')
def thankYouSignOut():
    global end_time, globalemail, logoutTime
    #if time is not up then render template
    message = calculateHours1(globalemail,logoutTime)
    if time.time()<=end_time:
        return render_template('thankyouSignOut.html', message = message)
    #javascript reloads the page every second so that the redirect can happen
    return redirect('/')

@app.route('/feedback', methods = ["get" ,"post"])
def feedBack():
    global end_time, globalemail
    globalemail = ""
    form = FeedBackForm()
    if form.is_submitted():
        result = request.form
        path = MasterPath + feedbackFile
        # write to csv file
        with open(path, 'a', newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow([result.get('feedback')])
        start_time = time.time()
        end_time = start_time + 1.0
        return redirect('/thankyou')
    return render_template('feedback.html', form=form)
class Person:
    def __init__(self, email, login_count):
        self.email = email
        self.login_count = login_count
    def __repr__(self):
            return f"{findName(self.email)}, logins: {self.login_count}"

    def __eq__(self, other):
        if isinstance(other, Person):
            return self.email == other.email
        return False
class PersonHours:
    def __init__(self, email, hours):
        self.email = email
        self.hours = hours

    def __repr__(self):
        return f"{findName(self.email)} | Hours: {self.hours:.2f}"

    def __eq__(self, other):
        if isinstance(other, PersonHours):
            return self.email == other.email
        return False
class Login:
    def __init__(self, email, login):
        self.email = email
        self.login = login

    def __repr__(self):
        return f"{self.email} logins={self.login}"
    def __eq__(self, other):
        if isinstance(other, Login):
            return self.email == other.email
        return False
class Logout:
    def __init__(self, email, logout):
        self.email = email
        self.logout = logout

    def __repr__(self):
        return f"{self.email}logouts={self.logout}"
    def __eq__(self, other):
        if isinstance(other, Logout):
            return self.email == other.email
        return False
@app.route('/leaderboard', methods = ["get" ,"post"])
def leaderBoard():
    global end_time, emailStat, sorted_by_hours
    try:
        form = SignInForm()
        message = ""
        if form.is_submitted():
            result = request.form
            recordTime = time.time()
            print(result)
            # Signin procedure
            # verification process
            inputEmail = result.get("email") + "@uccs.edu"
            pathR = MasterPath + emailList
            valid = False
            with open(pathR, 'r', newline="") as csvfile:
                reader = csv.reader(csvfile)
                for row in reader:
                    if row[0] == inputEmail:
                        valid = True
                        print(valid)
            if valid:
                # write to csv file
                emailStat = inputEmail
                # start timer for thank you page
                start_time = time.time()
                end_time = start_time + 1.0
                return redirect('/leaderboard/personalstats')
            else:
                message = "Invalid Email"
        Logins = []
        Logouts = []
        now = time.time()
        pathLogin = MasterPath + loginTimeFile
        pathLogout = MasterPath + logoutTimeFile
        with open(pathLogin, 'r', newline="") as csvfile:
            reader = csv.reader(csvfile)
            header = next(reader)
            for row in reader:
                # Logins within last 7 days
                Logins.append(Login(row[0], row[1]))
        with open(pathLogout, 'r', newline="") as csvfile:
            reader = csv.reader(csvfile)
            header = next(reader)
            for row in reader:
                Logouts.append(Logout(row[0], row[1]))
        sorted_by_hours = RankHours(Logins,Logouts)
        print(sorted_by_hours)
        students = []
        for i in range(50):
            hours=sorted_by_hours[i].hours
            special = ""
            emoji = ""
            if hours < 5:
                tier = "(New)"
            elif hours < 15:
                tier = "🪵"
            elif hours < 30:
                tier = "🟤"
            elif hours < 50:
                tier = "🔘"
            elif hours < 80:
                tier = "👑"
            elif hours < 120:
                tier = "🟩"
            elif hours < 160:
                tier = "♦️"
            else:
                tier = "💎"
            if checkTutor(sorted_by_hours[i].email):
                special = "- Tutor"
                emoji = "✏️"
            if sorted_by_hours[i].email == "jfauson@uccs.edu":
                special = "- Creator"
                emoji = "👾"
            students.append(emoji + " " + str(sorted_by_hours[i]) + " " + special + " " + tier)
        studentOfTheWeek = StudentOfTheWeek()
        
        #print(studentOfTheWeek)
    except:
        studentOfTheWeek = "N/A"
        students = ["N/A"]
    return render_template('leaderboard2.0.html', message=message,form=form, students=students, studentOfTheWeek=studentOfTheWeek)

@app.route('/leaderboard/personalstats')
def personalStats():
    global emailStat, sorted_by_hours
    loginCount=findLoginCount(emailStat)
    try:
        if checkTutor(emailStat):
            index = -1
            for j, s in enumerate(sorted_by_hours):
                if s.email == emailStat:
                    index = j
                    break
            Rank = str(index+1)+"- Tutor ✏️"
            hours = sorted_by_hours[index].hours
            if emailStat == 'jfauson@uccs.edu':
                Rank = str(index + 1) + "- Creator 👾"
        else:
            index = 0
            hours = 0
            for person in sorted_by_hours:
                if person.email == emailStat:
                    hours = person.hours
                    break
                index  += 1
            Rank = str(index + 1)
        if hours < 5:
            tier = "New"
        elif hours < 15:
            tier = "🪵Wood🪵"
        elif hours < 30:
            tier = "🟤Bronze🟤"
        elif hours < 50:
            tier = "🔘Silver🔘"
        elif hours < 80:
            tier = "👑Gold👑"
        elif hours < 120:
            tier = "🟩Emerald🟩"
        elif hours < 160:
            tier = "♦️Ruby♦️"
        else:
            tier = "💎Diamond💎"
    except ValueError:
        Rank = "N/A"
        tier = "New"
        loginCount = 0
    name = findName(emailStat)
    message1 = str(loginCount)
    message2 = calculateAllHours(emailStat)
    message3 = str(Rank)
    message4 = tier
    return render_template('leaderboardStat.html', message1=message1, message2=message2, message3=message3, message4=message4, name=name)

def StudentOfTheWeek():
    Logins = []
    Logouts = []
    now = time.time()
    pathLogin = MasterPath + loginTimeFile
    pathLogout = MasterPath + logoutTimeFile
    with open(pathLogin, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader)
        for row in reader:
            #Logins within last 7 days
            if float(row[1]) > now-(3600*24*7):
                Logins.append(Login(row[0],row[1]))
    with open(pathLogout, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader)
        for row in reader:
            if float(row[1]) > now-(3600*24*7):
                Logouts.append(Logout(row[0], row[1]))
    try:
        sorted_by_hours7 = RankHours(Logins, Logouts)
        print("Student of the week ranking:")
        print("-------------------")
        print(sorted_by_hours7)
        hours = 0
        for j, s in enumerate(sorted_by_hours):
            if s.email == sorted_by_hours7[0].email:
                index = j
                break
        if index != -1:
            hours = sorted_by_hours[index].hours
        if hours < 5:
            tier = "(New)"
        elif hours < 15:
            tier = "🪵"
        elif hours < 30:
            tier = "🟤"
        elif hours < 50:
            tier = "🔘"
        elif hours < 80:
            tier = "👑"
        elif hours < 120:
            tier = "🟩"
        elif hours < 160:
            tier = "♦️"
        else:
            tier = "💎"
        return str(sorted_by_hours7[0])+" " +tier
    except:
        return "No Student of the Week"
def findLoginCount(email):
    pathW = MasterPath + loginTimeFile
    count = 0
    with open(pathW, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if row[0] == email:
                count += 1
    return count


def calculateAllHours(student):
    Logins = []
    Logouts = []
    now = time.time()
    pathLogin = MasterPath + loginTimeFile
    pathLogout = MasterPath + logoutTimeFile
    with open(pathLogin, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader)
        for row in reader:
            #Logins within last 7 days
            if row[0] == student:
                Logins.append(Login(row[0],row[1]))
    with open(pathLogout, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader)
        for row in reader:
            if row[0] == student:
                Logouts.append(Logout(row[0], row[1]))
    Person1 = PersonHours(student, 0)
    for i in range(len(Logins)):
            hours = 0
            logoutFound = False
            for j in range(len(Logouts)):
                if Logins[i].email == Logouts[j].email:
                    if float(Logouts[j].logout) < float(Logins[i].login) + 3600*14 and float(Logouts[j].logout) > float(Logins[i].login):
                        hours = round((float(Logouts[j].logout)-float(Logins[i].login))/3600,2)
                        logoutFound = True
                        break
            if not logoutFound:
                hours = 1.0

            Person1.hours += hours
    return round(Person1.hours,2)
def calculateHours1(email, logoutTime):
    pathLogin = MasterPath + loginTimeFile
    pathLogout = MasterPath + logoutTimeFile
    loginList = []
    logoutList = []
    with open(pathLogin, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if email == row[0]:
                # if
                if float(row[1]) > logoutTime - 3600*14:
                    loginList.append(float(row[1]))
    with open(pathLogout, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if email == row[0]:
                # if
                if float(row[1]) > logoutTime - 3600*16:
                    logoutList.append(float(row[1]))
    if loginList:
        return round((float(logoutTime)-max(loginList))/3600,2)
    else:
        return 0
def checkRecent(email, loginTime):
    path = MasterPath + loginTimeFile
    with open(path, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if email == row[0]:
                #if
                if float(row[1]) > loginTime-3600/2:
                   return True
    return False
def checkTutor(email):
    path = MasterPath + '/Data/tutors.csv'
    tutor = False
    with open(path, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if row[0] == email:
                tutor = True
    return tutor
def findName(email):
    path = MasterPath + emailList
    with open(path, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if row[0] == email:
                if row[1]!= "":
                    fname = row[1]
                    lname = row[2]
                    return fname + " " + lname
    return email
def verifyEmail(result):
    inputEmail = result.get("email") + "@uccs.edu"
    pathR = MasterPath + emailList
    valid = False
    with open(pathR, 'r', newline="") as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if row[0] == inputEmail:
                valid = True
    return valid
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    global globalemail, end_time
    form = SignUpForm()
    message = ""
    globalemail = ""
    pathW = MasterPath + pendingListFile
    if form.is_submitted():
        result = request.form
        if not verifyEmail(result) and result.get("email") != '':
            inputEmail = result.get("email") + "@uccs.edu"
            fName = result.get('fName')
            lName = result.get('lName')
            with open(pathW, 'a', newline="") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow([inputEmail,fName,lName])
        end_time = time.time() + 1.0
        return redirect('/thankyou')
    return render_template('SignUp.html', form=form)
@app.route('/admin', methods=['GET', 'POST'])
def admin():
    global admin
    admin = False
    form = AdminForm()
    encrypted_password = b'gAAAAABo5eRedKCwMUWZ8BGCrOO4mXoR5tXtSsM_hHDW5OlDsCI2atMTrF_idHwBRKy-WBAYEjdRDO5VYTNTXm1W8R6QL-1Kkw=='
    key = b'KjQUVjHIjDzyMGAMzJhOkitWhmcGRrSlMH3K8S9iHAw='
    f = Fernet(key)
    decrypted_password = f.decrypt(encrypted_password)
    password = decrypted_password.decode()
    message = ""
    if form.is_submitted():
        result = request.form
        attempt = result.get('adminpass')
        if attempt == password:
            admin = True
            return redirect('/admin/verify')
        message = "Invalid password"
    return render_template('admin.html', form=form, message=message)
@app.route('/admin/verify', methods=['GET', 'POST'])
def admin_verify():
    emails = []
    form = VerifyForm()
    pathPending = MasterPath + pendingListFile
    pathEmail = MasterPath + emailList
    global admin
    # Read CSV file (assuming it's in /static/emails.csv)
    with open(pathPending, newline='') as f:
        reader = csv.reader(f)
        for row in reader:
            if row:  # skip empty lines
                emails.append(row)
    if request.method == 'POST':
        result = request.form
        selected_emails = request.form.getlist('selected_emails')
        if "Verify" in result:
            print("Email verified: ", selected_emails)
            with open(pathEmail, 'a', newline="") as csvfile:
                writer = csv.writer(csvfile)
                for item in selected_emails:
                    itemToAdd = item.strip("[]").split("', '")
                    itemToAdd[0] = itemToAdd[0].strip("'")
                    itemToAdd[2] = itemToAdd[2].strip("'")
                    writer.writerow(itemToAdd)
        with open(pathPending, 'w', newline="") as csvfile:
            writer = csv.writer(csvfile)
            for item1 in emails:

                remove = False
                for item2 in selected_emails:
                    if item2 == str(item1):
                        remove = True

                if not remove:
                    writer.writerow(item1)
        return redirect('/admin/verify')
    if admin == True:
        return render_template('adminverify.html', emails=emails, form=form)
    else:
        return redirect('/')
def RankHours(Logins,Logouts):
    PeopleByHours = []
    PeopleEmails = []
    for i in range(len(Logins)-1, -1, -1):
        hours = 0
        logoutFound = False
        nextlogin = -1
        j = i + 1
        while j < len(Logins) and float(Logins[i].login) + 3600 * 24 > float(Logins[j].login):
            if Logins[i].email == Logins[j].email:
                nextlogin = float(Logins[j].login)
            j += 1
        j = len(Logouts) - 1
        while Logins[i].login < Logouts[j].logout and j >= 0:
            j -= 1
        while j < len(Logouts) and float(Logins[i].login) + 3600 * 16 > float(Logouts[j].logout):

            if Logins[i].email == Logouts[j].email:
                if float(Logins[i].login) + 3600 * 14 > float(Logouts[j].logout) > float(Logins[i].login):
                    if float(Logouts[j].logout) < nextlogin or nextlogin == -1:
                        hours = round((float(Logouts[j].logout) - float(Logins[i].login)) / 3600, 2)
                        logoutFound = True
                        del Logouts[j]
                        break
            j += 1
        if not logoutFound:
            hours = 0.5
        if Logins[i].email not in PeopleEmails:
            PeopleByHours.append(PersonHours(Logins[i].email, hours))
            PeopleEmails.append(Logins[i].email)
        else:
            for person in PeopleByHours:
                if person.email == Logins[i].email:
                    person.hours += hours
                    break

    sorted_by_hours = sorted(PeopleByHours, key=lambda p: p.hours, reverse=True)
    return sorted_by_hours

if __name__ == '__main__':
    # === Main Logic ===
    #if not is_admin():
        #pyuac.runAsAdmin()
    LoadSettings()
    port = 5000 #Don't change
    url = f"http://127.0.0.1:{port}"

    # Use a thread to open the browser after a short delay to allow the server to start
    threading.Timer(1.25, lambda: webbrowser.open(url)).start()

    # Run the Flask application
    app.run(port=port, debug=True)